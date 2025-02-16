import openpyxl
from datetime import datetime
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

def basla():
    global stun
	
    try:
    	workbook = openpyxl.load_workbook(dosya_adı)
    	sheet = workbook.active
    except FileNotFoundError:
        messagebox.showwarning("Dikkat", "Dosya bulunamadı.\nLütfen tekrar dosya seçiniz.")
    except InvalidFileException:
        messagebox.showwarning("Dikkat", "Dosya formatı yanlış seçildi.\nLütfen tekrar dosya seçiniz.")

    # Tarih sütununu seç
    date_column = stun
    flag = True
    # Her hücreyi dolaş ve formatı değiştir
    for row in range(2, sheet.max_row + 1):  # İlk satır başlık olduğu için 2'den başla
        cell = sheet[f"{date_column}{row}"]
        try:
            # Tarih değerini datetime nesnesine dönüştür
            date_value = datetime.strptime(str(cell.value), '%Y%m%d')
            # Yeni formatı uygula
            cell.value = date_value.strftime('%d.%m.%Y')    
        except ValueError:
            # Tarih formatı uygun değilse, hücreyi olduğu gibi bırak
            flag = False
            messagebox.showwarning("Dikkat", "Tarih sütunu yanlış seçildi veya sütun formata uygun değildir.\nLütfen tekrar sütun seçiniz.")
            break
    if flag:
        # Excel dosyasını kaydet
        workbook.save(f"{dosya_adı.split('.')[0]}_guncellendi.xlsx")
        messagebox.showinfo("Bilgi", f"Tarih güncellendi.Dosya {dosya_adı.split('.')[0]}_guncellendi.xlsx olarak kaydedildi.")
        buton_guncel.config(text="Güncellendi", bg="blue")

def stun_secilen(*args):
    global stun
    
    stun = label_option.get()

    if buton_guncel["text"] == "Güncellendi":
        buton_guncel.config(text="Güncelle", bg="#F0F0F0")
        

def dosyaac():
    global dosya_adı, stun
    
    try:
        dosya_adı = filedialog.askopenfilename(initialdir = "~/Downloads", title = "Dosya seçiniz", filetypes = (("Excel dosyaları", "*.xls*;*.xlsx*;*.xlsm*;*.xlsb*;*.csv*"),
                                                                                                   ("Tüm dosyalar", "*.*")))
    except FileNotFoundError:
        messagebox.showwarning("Dikkat", "Dosya bulunamadı.\nLütfen tekrar dosya seçiniz.")
    except InvalidFileException:
        messagebox.showwarning("Dikkat", "Dosya formatı yanlış seçildi.\nLütfen tekrar dosya seçiniz.")

    if dosya_adı == "":
        buton_dosya.config(text="Dosya seç", bg="red")
        messagebox.showwarning("Dikkat", "Dosya seçilmedi.\nLütfen tekrar dosya seçiniz.")
        
    if dosya_adı and dosya_adı != "":
        buton_dosya.config(text="Dosya seçildi", bg="blue")
    
    if buton_guncel["text"] == "Güncellendi":
        buton_guncel.config(text="Güncelle", bg="#F0F0F0")
        
    try:
    	workbook = openpyxl.load_workbook(dosya_adı)
    	sheet = workbook.active
    except FileNotFoundError:
        messagebox.showwarning("Dikkat", "Dosya bulunamadı.\nLütfen tekrar dosya seçiniz.")
    except InvalidFileException:
        messagebox.showwarning("Dikkat", "Dosya formatı yanlış seçildi.\nLütfen tekrar dosya seçiniz.")

    stunlar = []
    
    for column in range(1, sheet.max_column + 1):
        column_letter = openpyxl.utils.get_column_letter(column)
        column_name = sheet.cell(row=1, column=column).value
        if column_name is not None:
            stunlar.append(column_letter)
            
    label_open_menu = OptionMenu(main, label_option, *stunlar, command=stun_secilen) .place(x=300, y=200)
    stun = label_option.get()


#Gorsel Arayuz
main = Tk()
main.title("Tarih Format Guncelleme")
main.geometry("600x400")

main.resizable(True, True)
main.bind("<Escape>", lambda event: main.destroy())
main.bind("<Control_L>", lambda event: dosyaac())
main.bind("<Return>", lambda event: basla())

Label(main, text = "Güncellenecek dosyayı seçiniz :").place(x=50, y=100)
label_option = StringVar(main)
label_option.set("A")

try:
    buton_dosya = Button(
        main,
        text="Dosya Seç",
        padx=5,
        pady=5,
        bd=5,
        activebackground="yellow",
        relief=GROOVE,
        command=dosyaac,
    )
    buton_dosya.place(x=300, y=100)
except FileNotFoundError:
    messagebox.showwarning("Dikkat", "Dosya bulunamadı.\nLütfen tekrar dosya seçiniz.")
except InvalidFileException:
        messagebox.showwarning("Dikkat", "Dosya formatı yanlış seçildi.\nLütfen tekrar dosya seçiniz.")

Label(main, text = "Sütunu seçiniz :").place(x=50, y=200)

buton_guncel = Button(
main,
    text="Güncelle",
    padx=5,
    pady=5,
    bd=5,
    activebackground="yellow",
    relief=GROOVE,
    command=basla,
)
buton_guncel.place(x=300, y=300)

main.mainloop()